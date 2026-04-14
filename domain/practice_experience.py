"""Structured admin-governed operational experience imported from practice sheets."""

from __future__ import annotations

from collections import Counter
from datetime import date, datetime
import hashlib
import math
import re
import statistics
from typing import Any, Iterable
import unicodedata

from core.validators import validate_operational_feedback_status
from domain.berth_layout import canonicalize_berth_label, is_anchorage_berth, is_known_berth_label
from domain.document_processing import iso_now


PRACTICE_EXPERIENCE_STATE_KEY = "practice_maneuver_experience"
PRACTICE_EXPERIENCE_VERSION = 1
PRACTICE_EXPERIENCE_ACTIVE_STATUSES = {"approved", "avoid"}

_MANEUVER_TYPE_ALIASES = {
    "entrada": ("entry", "Entrada"),
    "entry": ("entry", "Entrada"),
    "saida": ("departure", "Saída"),
    "saída": ("departure", "Saída"),
    "departure": ("departure", "Saída"),
    "mudanca": ("shift", "Mudança"),
    "mudança": ("shift", "Mudança"),
    "shift": ("shift", "Mudança"),
    "fundear": ("shift", "Fundear"),
}


def _clean_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _as_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    clean = _clean_text(value).replace(" ", "").replace(",", ".")
    if not clean:
        return None
    try:
        return float(clean)
    except ValueError:
        return None


def _duration_hours(value: Any) -> float | None:
    if value is None:
        return None
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return float(value.hour) + float(value.minute) / 60 + float(getattr(value, "second", 0)) / 3600
    number = _as_number(value)
    if number is None:
        return None
    return number * 24 if 0 < number < 1 else number


def _split_length_beam(value: Any) -> tuple[float | None, float | None]:
    text = _clean_text(value)
    if not text:
        return None, None

    parts = re.findall(r"\d+(?:[.,]\d+)?", text)
    if len(parts) < 2:
        return None, None
    return _as_number(parts[0]), _as_number(parts[1])


def _median(values: Iterable[float | None]) -> float | None:
    clean = [value for value in values if value is not None]
    return statistics.median(clean) if clean else None


def _mean(values: Iterable[float | None]) -> float | None:
    clean = [value for value in values if value is not None]
    return statistics.mean(clean) if clean else None


def _format_metric(value: float | None, *, decimals: int = 1, suffix: str = "") -> str:
    if value is None:
        return "--"
    formatted = f"{value:.{decimals}f}".rstrip("0").rstrip(".")
    return f"{formatted}{suffix}"


def _band(value: float | None, width: int, suffix: str = "m") -> str:
    if value is None:
        return "sem registo"
    start = math.floor(value / width) * width
    end = start + width
    return f"{start}-{end:g}{suffix}"


def _case_key(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", _clean_text(value).lower())
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", ascii_value).strip()


def _normalize_maneuver_type(value: Any) -> tuple[str, str]:
    clean = _clean_text(value).lower()
    return _MANEUVER_TYPE_ALIASES.get(clean, (clean, _clean_text(value)))


def _date_label(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _clean_text(value)


def _status_meta(value: str) -> tuple[str, str]:
    clean = (value or "").strip().lower()
    if clean == "approved":
        return "Aprovado", "online"
    if clean == "avoid":
        return "Evitar", "degraded"
    if clean == "review":
        return "Rever", "degraded"
    return "Por rever", "neutral"


def _operational_route(maneuver_type: str, berth: str) -> tuple[str, str]:
    if maneuver_type == "entry":
        return "Fora", berth
    if maneuver_type == "departure":
        return berth, "Fora"
    if maneuver_type == "shift":
        return "", berth
    return "", berth


def _decision_flags_from_comments(comments: Iterable[str], maneuver_type: str) -> list[str]:
    text = " ".join(comment.lower() for comment in comments if comment)
    flags: list[str] = []
    if any(token in text for token in ("ondul", "vaga", "mar grosso", "agitação", "agitacao")):
        flags.append("wave_related")
    if any(token in text for token in ("suspens", "suspensa", "suspender")):
        flags.append("pilotage_suspended")
    if any(token in text for token in ("cancel", "cancelada", "cancelado")):
        flags.append("pilotage_cancelled")
    if maneuver_type == "entry" and "wave_related" in flags:
        flags.append("entry_aborted_by_sea_state")
    return flags


def _experience_id(key_parts: Iterable[Any]) -> str:
    digest = hashlib.sha1("|".join(_clean_text(value) for value in key_parts).encode("utf-8")).hexdigest()[:12]
    return f"practice-{digest}"


def _row_payload(raw_headers: list[str], raw_row: tuple[Any, ...]) -> dict[str, Any]:
    payload = dict(zip(raw_headers, raw_row))
    maneuver_type, maneuver_type_label = _normalize_maneuver_type(payload.get("Tipo Manobra"))
    vessel_name = _clean_text(payload.get("Nome"))
    vessel_type = _clean_text(payload.get("Tipo Navio")) or "Não indicado"
    berth = canonicalize_berth_label(payload.get("Cais")) or _clean_text(payload.get("Cais")) or "Não indicado"
    length_m, beam_m = _split_length_beam(payload.get("Dimensão C/B (m)"))
    return {
        "source_row": _clean_text(payload.get("No")),
        "date": payload.get("Data"),
        "date_label": _date_label(payload.get("Data")),
        "maneuver_type": maneuver_type,
        "maneuver_type_label": maneuver_type_label,
        "vessel_name": vessel_name,
        "vessel_type": vessel_type,
        "gt": _as_number(payload.get("GT")),
        "length_m": length_m,
        "beam_m": beam_m,
        "draft_m": _as_number(payload.get("Calado (m)")),
        "tug_count": _as_number(payload.get("Rebocadores")),
        "berth": berth,
        "duration_h": _duration_hours(payload.get("Tempo de Manobra (h)")),
        "comment": _clean_text(payload.get("Comentários")),
    }


def _record_from_bucket(
    *,
    key: tuple[str, str, str, str, str, str],
    rows: list[dict[str, Any]],
    source_filename: str,
    feedback_status: str,
    imported_by: str,
    imported_at: str,
) -> dict:
    maneuver_type, vessel_type_key, _berth_key, loa_band, beam_band, draft_band = key
    maneuver_label = rows[0]["maneuver_type_label"]
    vessel_type = rows[0]["vessel_type"]
    berth = rows[0]["berth"]
    origin, destination = _operational_route(maneuver_type, berth)
    duration_median = _median(row["duration_h"] for row in rows)
    duration_mean = _mean(row["duration_h"] for row in rows)
    tug_counter = Counter(
        str(int(row["tug_count"])) if row["tug_count"] is not None else "sem registo"
        for row in rows
    )
    tug_value = tug_counter.most_common(1)[0][0] if tug_counter else ""
    tug_for_feature = "" if tug_value == "sem registo" else tug_value
    comments = [row["comment"] for row in rows if row.get("comment")]
    vessel_counter = Counter(row["vessel_name"] for row in rows if row.get("vessel_name"))
    vessel_examples = [name for name, _count in vessel_counter.most_common(5)]
    length_median = _median(row["length_m"] for row in rows)
    beam_median = _median(row["beam_m"] for row in rows)
    draft_median = _median(row["draft_m"] for row in rows)
    gt_median = _median(row["gt"] for row in rows)
    date_labels = [row["date_label"] for row in rows if row.get("date_label")]
    date_range = ""
    if date_labels:
        date_range = f"{min(date_labels)} a {max(date_labels)}"
    record_id = _experience_id([source_filename, *key])
    summary = (
        f"{maneuver_label} | {vessel_type} | {berth} | {loa_band} | "
        f"{len(rows)} caso(s) | rebocadores mais comuns {tug_value} | "
        f"duração mediana {_format_metric(duration_median, suffix=' h')}"
    )
    return {
        "id": record_id,
        "maneuver_id": record_id,
        "port_call_id": "",
        "reference_code": f"EXP-{record_id[-6:].upper()}",
        "source_type": "practice_import",
        "source_label": "Experiência prática importada",
        "source_filename": source_filename,
        "source_rows": [row["source_row"] for row in rows if row.get("source_row")][:20],
        "vessel_name": f"Padrão {vessel_type} · {berth}",
        "maneuver_type": maneuver_type,
        "maneuver_type_label": maneuver_label,
        "current_state": "completed",
        "current_state_label": "Realizada",
        "origin_label": origin,
        "destination_label": destination,
        "planned_at": None,
        "decided_at": None,
        "completed_at": None,
        "reported_at": None,
        "latest_event_at": imported_at,
        "case_summary": summary,
        "practice_summary": summary,
        "practice_metrics": {
            "case_count": len(rows),
            "date_range": date_range,
            "duration_median_h": duration_median,
            "duration_mean_h": duration_mean,
            "duration_median_label": _format_metric(duration_median, suffix=" h"),
            "duration_mean_label": _format_metric(duration_mean, suffix=" h"),
            "dominant_tug_count": tug_value,
            "tug_distribution_label": ", ".join(f"{label} ({count})" for label, count in tug_counter.most_common(5)),
            "vessel_examples": vessel_examples,
            "comments_count": len(comments),
            "comments": comments[:5],
            "loa_band": loa_band,
            "beam_band": beam_band,
            "draft_band": draft_band,
        },
        "vessel_snapshot": {
            "type": vessel_type,
            "loa_m": _format_metric(length_median, suffix=""),
            "beam_m": _format_metric(beam_median, suffix=""),
            "gt_t": _format_metric(gt_median, decimals=0),
            "max_draft_m": _format_metric(draft_median, suffix=""),
            "examples": vessel_examples,
        },
        "scale_snapshot": {
            "source": source_filename,
            "date_range": date_range,
            "notes": "Dados agregados de prática importada; a folha não inclui meteorologia nem IMO.",
        },
        "planning_snapshot": {
            "origin": origin,
            "destination": destination,
            "planned_draft_m": _format_metric(draft_median, suffix=""),
            "tug_count": tug_for_feature,
            "plan_note": summary,
            "plan_observations": " | ".join(comments[:3]),
            "created_by": imported_by,
            "created_at": imported_at,
        },
        "decision_snapshot": {
            "decision": "approved",
            "state": "completed",
            "approval_note": "Padrão agregado de experiência prática importada pelo admin.",
            "decided_by": imported_by,
            "decided_at": imported_at,
        },
        "execution_snapshot": {
            "report_note": " | ".join(comments[:3]),
            "reported_by": "experiência importada",
            "reported_at": imported_at,
        },
        "outcome_snapshot": {
            "state": "completed",
            "state_label": "Realizada",
            "report_completed": bool(comments),
            "decision_flags": _decision_flags_from_comments(comments, maneuver_type),
        },
        "environment_snapshot": {
            "latest": {
                "status": "not_captured",
                "source": "practice_import",
                "reason": "A folha de prática não inclui meteorologia nem leitura costeira.",
            }
        },
        "feature_snapshot": {
            "maneuver_type": maneuver_type,
            "origin": origin,
            "destination": destination,
            "origin_key": _case_key(canonicalize_berth_label(origin) or origin),
            "destination_key": _case_key(canonicalize_berth_label(destination) or destination),
            "origin_is_anchorage": is_anchorage_berth(origin),
            "destination_is_anchorage": is_anchorage_berth(destination),
            "origin_is_known_berth": is_known_berth_label(origin),
            "destination_is_known_berth": is_known_berth_label(destination),
            "vessel_type": vessel_type,
            "vessel_type_key": vessel_type_key,
            "vessel_loa_m": length_median,
            "vessel_beam_m": beam_median,
            "vessel_gt_t": gt_median,
            "planned_draft_m": draft_median,
            "reported_draft_m": draft_median,
            "bow_thruster": "unknown",
            "stern_thruster": "unknown",
            "tug_count": tug_for_feature,
            "constraints": [],
            "wave_sensitive": maneuver_type in {"entry", "departure"},
        },
        "change_log": [],
        "feedback_status": feedback_status,
        "feedback_note": "Importado e aprovado pelo admin." if feedback_status == "approved" else "",
        "feedback_updated_by": imported_by,
        "feedback_updated_at": imported_at,
        "created_at": imported_at,
        "updated_at": imported_at,
    }


def build_practice_experience_records_from_xlsx(
    source: Any,
    *,
    source_filename: str,
    imported_by: str,
    feedback_status: str = "approved",
) -> tuple[list[dict], dict]:
    """Parse and aggregate the practice spreadsheet into governed maneuver patterns."""
    feedback_status = validate_operational_feedback_status(feedback_status)
    try:
        import openpyxl
    except ImportError as exc:
        raise ValueError("A dependência openpyxl não está instalada para ler ficheiros .xlsx.") from exc

    workbook = openpyxl.load_workbook(source, data_only=True, read_only=True)
    if "Dados" not in workbook.sheetnames:
        raise ValueError("O ficheiro tem de conter a folha 'Dados'.")
    worksheet = workbook["Dados"]
    headers = [_clean_text(value) for value in next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))]

    rows: list[dict[str, Any]] = []
    skipped = 0
    for raw_row in worksheet.iter_rows(min_row=3, values_only=True):
        row = _row_payload(headers, raw_row)
        if not row["maneuver_type"] or not row["vessel_name"] or row["maneuver_type"] not in {"entry", "departure", "shift"}:
            skipped += 1
            continue
        rows.append(row)
    if not rows:
        raise ValueError("Não encontrei manobras válidas na folha 'Dados'.")

    buckets: dict[tuple[str, str, str, str, str, str], list[dict[str, Any]]] = {}
    for row in rows:
        key = (
            row["maneuver_type"],
            _case_key(row["vessel_type"]),
            _case_key(row["berth"]),
            _band(row["length_m"], 50),
            _band(row["beam_m"], 5),
            _band(row["draft_m"], 2),
        )
        buckets.setdefault(key, []).append(row)

    imported_at = iso_now()
    records = [
        _record_from_bucket(
            key=key,
            rows=bucket,
            source_filename=source_filename,
            feedback_status=feedback_status,
            imported_by=imported_by,
            imported_at=imported_at,
        )
        for key, bucket in buckets.items()
    ]
    records.sort(
        key=lambda item: (
            item.get("maneuver_type", ""),
            -int((item.get("practice_metrics") or {}).get("case_count") or 0),
            item.get("vessel_name", ""),
        )
    )
    type_counter = Counter(row["maneuver_type_label"] for row in rows)
    return records, {
        "source_filename": source_filename,
        "raw_rows": len(rows),
        "skipped_rows": skipped,
        "pattern_count": len(records),
        "comments_count": sum(1 for row in rows if row.get("comment")),
        "maneuver_types_label": ", ".join(f"{label} ({count})" for label, count in type_counter.most_common()),
    }


def practice_experience_state(store: Any) -> dict:
    state = store.get_runtime_state(PRACTICE_EXPERIENCE_STATE_KEY) or {}
    records = state.get("records") or []
    return {
        "version": PRACTICE_EXPERIENCE_VERSION,
        "records": [item for item in records if isinstance(item, dict)],
        "sources": state.get("sources") or {},
        "updated_at": state.get("updated_at") or "",
        "updated_by": state.get("updated_by") or "",
    }


def list_practice_experience_records(
    store: Any,
    *,
    feedback_statuses: set[str] | None = None,
    limit: int | None = None,
) -> list[dict]:
    records = list(practice_experience_state(store)["records"])
    if feedback_statuses is not None:
        allowed = {(value or "").strip().lower() for value in feedback_statuses}
        records = [item for item in records if (item.get("feedback_status") or "").strip().lower() in allowed]
    records.sort(
        key=lambda item: (
            item.get("updated_at") or item.get("created_at") or "",
            int((item.get("practice_metrics") or {}).get("case_count") or 0),
        ),
        reverse=True,
    )
    decorated = []
    for item in records:
        label, badge = _status_meta(item.get("feedback_status", ""))
        metrics = item.get("practice_metrics") or {}
        decorated.append(
            {
                **item,
                "feedback_status_label": label,
                "feedback_badge": badge,
                "case_count": int(metrics.get("case_count") or 0),
                "duration_median_label": metrics.get("duration_median_label") or "--",
                "tug_distribution_label": metrics.get("tug_distribution_label") or "--",
                "vessel_examples_label": ", ".join(metrics.get("vessel_examples") or []) or "--",
                "comments_label": " | ".join(metrics.get("comments") or []) or "--",
                "profile_label": (
                    f"{(item.get('vessel_snapshot') or {}).get('type') or '--'} · "
                    f"LOA {metrics.get('loa_band') or '--'} · "
                    f"Boca {metrics.get('beam_band') or '--'} · "
                    f"Calado {metrics.get('draft_band') or '--'}"
                ),
                "route_label": f"{item.get('origin_label') or '--'} → {item.get('destination_label') or '--'}",
            }
        )
    if limit is not None:
        return decorated[: max(limit, 0)]
    return decorated


def save_practice_experience_records(
    store: Any,
    records: list[dict],
    *,
    source_filename: str,
    updated_by: str,
    replace_source: bool = True,
) -> dict:
    state = practice_experience_state(store)
    existing = list(state["records"])
    if replace_source:
        existing = [item for item in existing if item.get("source_filename") != source_filename]
    by_id = {item.get("id"): item for item in existing if item.get("id")}
    for record in records:
        by_id[record["id"]] = record
    next_records = list(by_id.values())
    now = iso_now()
    sources = dict(state.get("sources") or {})
    sources[source_filename] = {
        "record_count": len(records),
        "updated_by": updated_by,
        "updated_at": now,
    }
    payload = {
        "version": PRACTICE_EXPERIENCE_VERSION,
        "records": next_records,
        "sources": sources,
        "updated_at": now,
        "updated_by": updated_by,
    }
    store.set_runtime_state(PRACTICE_EXPERIENCE_STATE_KEY, payload)
    return payload


def update_practice_experience_feedback(
    store: Any,
    record_id: str,
    *,
    feedback_status: str,
    feedback_note: str = "",
    feedback_by: str = "",
) -> dict:
    feedback_status = validate_operational_feedback_status(feedback_status)
    state = practice_experience_state(store)
    now = iso_now()
    updated = None
    for record in state["records"]:
        if record.get("id") != record_id:
            continue
        record["feedback_status"] = feedback_status
        record["feedback_note"] = _clean_text(feedback_note)
        record["feedback_updated_by"] = _clean_text(feedback_by)
        record["feedback_updated_at"] = now
        record["updated_at"] = now
        updated = record
        break
    if not updated:
        raise ValueError("Experiência prática não encontrada.")
    state["updated_at"] = now
    state["updated_by"] = _clean_text(feedback_by)
    store.set_runtime_state(PRACTICE_EXPERIENCE_STATE_KEY, state)
    return updated


def delete_practice_experience_record(store: Any, record_id: str, *, deleted_by: str = "") -> int:
    state = practice_experience_state(store)
    records = [item for item in state["records"] if item.get("id") != record_id]
    removed = len(state["records"]) - len(records)
    if not removed:
        return 0
    state["records"] = records
    state["updated_at"] = iso_now()
    state["updated_by"] = _clean_text(deleted_by)
    store.set_runtime_state(PRACTICE_EXPERIENCE_STATE_KEY, state)
    return removed


def clear_practice_experience_records(store: Any, *, cleared_by: str = "") -> int:
    state = practice_experience_state(store)
    removed = len(state["records"])
    store.set_runtime_state(
        PRACTICE_EXPERIENCE_STATE_KEY,
        {
            "version": PRACTICE_EXPERIENCE_VERSION,
            "records": [],
            "sources": {},
            "updated_at": iso_now(),
            "updated_by": _clean_text(cleared_by),
        },
    )
    return removed
