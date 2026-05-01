#!/usr/bin/env python3
"""Generate structured practice maneuver experience from the local workbook."""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, time, timezone, timedelta
from decimal import Decimal, InvalidOperation
import argparse
import hashlib
import json
import math
from pathlib import Path
import re
import statistics
import sys
from typing import Any
import unicodedata

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from domain.berth_layout import (  # noqa: E402
    canonicalize_berth_label,
    is_anchorage_berth,
    is_known_berth_label,
)

KIND = "pragtico.practice_maneuver_experience"
VERSION = 1
DEFAULT_INPUT = PROJECT_ROOT / "review" / "Manobras Pratica.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "knowledge" / "practice_maneuver_experience.json"

LABEL_ENTRY = "Entrada"
LABEL_DEPARTURE = "Sa\u00edda"
LABEL_SHIFT = "Mudan\u00e7a"
LABEL_ANCHOR = "Fundear"
LABEL_UNKNOWN = "N\u00e3o indicado"

TYPE_TO_CODE = {
    LABEL_ENTRY: "entry",
    LABEL_DEPARTURE: "departure",
    LABEL_SHIFT: "shift",
    LABEL_ANCHOR: "shift",
}

TYPE_ORDER = {
    LABEL_DEPARTURE: 0,
    LABEL_ENTRY: 1,
    LABEL_SHIFT: 2,
    LABEL_ANCHOR: 3,
}

SLASH_ALIAS_KEYS = {
    "secil outao w",
    "secil outao e",
    "praias do sado pirites alentejanas",
    "fundeadouro sul troia",
    "cais 10 autoeuropa",
    "cais 11 autoeuropa",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def key_text(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", clean_text(value).lower())
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", ascii_value).strip()


def compact_key(value: Any) -> str:
    return key_text(value).replace(" ", "")


def parse_decimal(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value).replace(",", ".")
    try:
        return float(Decimal(text))
    except InvalidOperation:
        pass
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def format_number(value: float | int | None, *, decimals: int = 1) -> str:
    if value is None:
        return "--"
    number = float(value)
    if number.is_integer():
        return str(int(number))
    text = f"{number:.{decimals}f}".rstrip("0").rstrip(".")
    return text or "0"


def parse_gt(value: Any) -> float | None:
    text = clean_text(value).lower().replace(",", ".")
    match = re.fullmatch(r"(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)", text)
    if match:
        return float(match.group(1)) * float(match.group(2))
    return parse_decimal(value)


def parse_dimensions(value: Any) -> tuple[float | None, float | None]:
    text = clean_text(value).lower().replace(",", ".")
    match = re.fullmatch(
        r"(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)",
        text,
    )
    if match:
        units = float(match.group(1))
        loa = units * float(match.group(2))
        beam = float(match.group(3))
        return loa, beam
    parts = re.split(r"\s*/\s*", text)
    loa = parse_decimal(parts[0]) if parts else None
    beam = parse_decimal(parts[1]) if len(parts) > 1 else None
    return loa, beam


def duration_hours(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, timedelta):
        return value.total_seconds() / 3600
    if isinstance(value, datetime):
        return value.hour + value.minute / 60 + value.second / 3600
    if isinstance(value, time):
        return value.hour + value.minute / 60 + value.second / 3600
    if isinstance(value, (int, float)):
        return float(value) * 24 if 0 <= float(value) < 1 else float(value)
    return None


def date_label(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean_text(value)


def band(value: float | None, step: int) -> str:
    if value is None:
        return "sem registo"
    low = math.floor(value / step) * step
    return f"{int(low)}-{int(low + step)}m"


def label_hours(value: float | None) -> str:
    if value is None:
        return "--"
    rounded = round(float(value), 1)
    if rounded.is_integer():
        return f"{int(rounded)} h"
    return f"{rounded:.1f} h"


def endpoint_alias(value: Any) -> str:
    raw = clean_text(value)
    if not raw:
        return ""
    key = key_text(raw)
    compact = key.replace(" ", "")
    aliases = {
        "auto europa": "Auto Europa",
        "autoeuropa": "Auto Europa",
        "eco oil": "Eco-Oil (lado montante)",
        "ecooil": "Eco-Oil (lado montante)",
        "ecoil": "Eco-Oil (lado montante)",
        "eco oil lado montante": "Eco-Oil (lado montante)",
        "tanquisado": "Tanquisado (lado jusante)",
        "tanquisado lado jusante": "Tanquisado (lado jusante)",
        "praias do sado": "Praias do Sado / Pirites Alentejanas",
        "pirites": "Praias do Sado / Pirites Alentejanas",
        "pirites alentejanas": "Praias do Sado / Pirites Alentejanas",
        "sapec liq": "SAPEC L\u00edquidos",
        "sapec liquidos": "SAPEC L\u00edquidos",
        "tgl": "SAPEC L\u00edquidos",
        "sapec s": "SAPEC S\u00f3lidos",
        "sapec s.": "SAPEC S\u00f3lidos",
        "sapec solidos": "SAPEC S\u00f3lidos",
        "tps": "SAPEC S\u00f3lidos",
        "f n": "Fundeadouro Norte",
        "f n.": "Fundeadouro Norte",
        "f.n": "Fundeadouro Norte",
        "f.n.": "Fundeadouro Norte",
        "fund n": "Fundeadouro Norte",
        "fund. norte": "Fundeadouro Norte",
        "fund norte": "Fundeadouro Norte",
        "fundeadouro n": "Fundeadouro Norte",
        "fundeadouro norte": "Fundeadouro Norte",
        "f s": "Fundeadouro Sul / Tr\u00f3ia",
        "f s.": "Fundeadouro Sul / Tr\u00f3ia",
        "f.s": "Fundeadouro Sul / Tr\u00f3ia",
        "f.s.": "Fundeadouro Sul / Tr\u00f3ia",
        "f troia": "Fundeadouro Sul / Tr\u00f3ia",
        "f. troia": "Fundeadouro Sul / Tr\u00f3ia",
        "fund troia": "Fundeadouro Sul / Tr\u00f3ia",
        "fund. troia": "Fundeadouro Sul / Tr\u00f3ia",
        "troia": "Fundeadouro Sul / Tr\u00f3ia",
        "fundeadouro s": "Fundeadouro Sul / Tr\u00f3ia",
        "fundeadouro sul": "Fundeadouro Sul / Tr\u00f3ia",
        "c10": "Cais 10 / Autoeuropa",
        "c11": "Cais 11 / Autoeuropa",
        "cais 10": "Cais 10 / Autoeuropa",
        "cais 11": "Cais 11 / Autoeuropa",
        "tms2": "TMS 2",
        "tms 2": "TMS 2",
        "tms1": "TMS 1",
        "tms 1": "TMS 1",
    }
    if key in aliases:
        return aliases[key]
    if compact in aliases:
        return aliases[compact]
    canonical = canonicalize_berth_label(raw)
    return canonical or raw


def is_route_label(value: str, maneuver_label: str) -> bool:
    if maneuver_label != LABEL_SHIFT or "/" not in value:
        return False
    return key_text(value) not in SLASH_ALIAS_KEYS


def route_from_row(maneuver_label: str, berth_value: Any) -> tuple[str, str, str]:
    berth = clean_text(berth_value)
    if maneuver_label == LABEL_ENTRY:
        destination = endpoint_alias(berth)
        return "Fora", destination, destination
    if maneuver_label == LABEL_DEPARTURE:
        origin = endpoint_alias(berth)
        return origin, "Fora", origin
    if is_route_label(berth, maneuver_label):
        left, right = [part.strip() for part in berth.split("/", 1)]
        origin = endpoint_alias(left)
        destination = endpoint_alias(right)
        route = f"{origin} -> {destination}" if origin and destination else endpoint_alias(berth)
        return origin, destination, route
    destination = endpoint_alias(berth)
    return "", destination, destination


def tug_label(value: Any) -> str:
    number = parse_decimal(value)
    if number is None:
        return ""
    if number.is_integer():
        return str(int(number))
    return str(int(round(number)))


def row_is_usable(values: list[Any]) -> bool:
    no = values[0]
    if not isinstance(no, (int, float)):
        return False
    if not clean_text(values[5]) or not clean_text(values[6]):
        return False
    if values[11] is None or clean_text(values[11]) == "":
        return False
    return True


def parse_rows(path: Path, sheet_name: str) -> tuple[list[dict[str, Any]], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows: list[dict[str, Any]] = []
    skipped = 0
    for excel_row, row in enumerate(sheet.iter_rows(min_row=3, max_col=15, values_only=True), start=3):
        values = list(row)
        if not any(value is not None for value in values):
            continue
        has_operational_content = any(values[index] is not None for index in [1, 5, 6, 7, 8, 9, 10, 11, 12, 14])
        if not row_is_usable(values):
            if has_operational_content:
                skipped += 1
            continue
        loa, beam = parse_dimensions(values[9])
        maneuver_label = clean_text(values[5])
        origin, destination, route = route_from_row(maneuver_label, values[12])
        rows.append(
            {
                "no": int(values[0]),
                "excel_row": excel_row,
                "date": values[1],
                "date_label": date_label(values[1]),
                "duration_h": duration_hours(values[4]),
                "maneuver_label": maneuver_label,
                "maneuver_type": TYPE_TO_CODE.get(maneuver_label, "shift"),
                "vessel_name": clean_text(values[6]),
                "vessel_type": clean_text(values[7]) or LABEL_UNKNOWN,
                "gt": parse_gt(values[8]),
                "loa": loa,
                "beam": beam,
                "draft": parse_decimal(values[10]),
                "tugs": tug_label(values[11]),
                "berth": clean_text(values[12]),
                "origin": origin,
                "destination": destination,
                "route": route,
                "comment": clean_text(values[14]),
                "loa_band": band(loa, 50),
                "beam_band": band(beam, 5),
                "draft_band": band(parse_decimal(values[10]), 2),
            }
        )
    return rows, skipped


def group_key(row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        row["maneuver_label"],
        row["vessel_type"],
        row["origin"],
        row["destination"],
        row["route"],
        row["loa_band"],
        row["beam_band"],
        row["draft_band"],
    )


def median(values: list[float | None]) -> float | None:
    clean_values = [float(value) for value in values if value is not None]
    return float(statistics.median(clean_values)) if clean_values else None


def mean(values: list[float | None]) -> float | None:
    clean_values = [float(value) for value in values if value is not None]
    return float(statistics.mean(clean_values)) if clean_values else None


def first_distinct(values: list[str], limit: int = 5) -> list[str]:
    output: list[str] = []
    seen: set[str] = set()
    for value in values:
        clean = clean_text(value)
        if not clean:
            continue
        marker = key_text(clean)
        if marker in seen:
            continue
        seen.add(marker)
        output.append(clean)
        if len(output) >= limit:
            break
    return output


def dominant(values: list[str]) -> str:
    counter = Counter(values)
    return counter.most_common(1)[0][0] if counter else "sem registo"


def distribution_label(values: list[str]) -> str:
    counter = Counter(values)
    if not counter:
        return "sem registo"
    return ", ".join(f"{label} ({count})" for label, count in counter.most_common())


def case_id(group: tuple[Any, ...]) -> str:
    digest = hashlib.sha1(json.dumps(group, ensure_ascii=False, sort_keys=False).encode("utf-8")).hexdigest()[:12]
    return f"practice-{digest}"


def is_wave_sensitive(vessel_type: str) -> bool:
    key = key_text(vessel_type)
    return any(marker in key for marker in ("contentores", "roro", "ferry", "high speed craft"))


def build_record(group: tuple[Any, ...], rows: list[dict[str, Any]], generated_at: str) -> dict[str, Any]:
    rows = sorted(rows, key=lambda item: item["no"])
    first = rows[0]
    record_id = case_id(group)
    durations = [row["duration_h"] for row in rows]
    duration_median = median(durations)
    duration_mean = mean(durations)
    tug_values = [row["tugs"] for row in rows if row["tugs"]]
    dominant_tug = dominant(tug_values)
    comments = [row["comment"] for row in rows if row["comment"]]
    examples = first_distinct([row["vessel_name"] for row in rows])
    dates = [row["date_label"] for row in rows if row["date_label"]]
    date_range = f"{min(dates)} a {max(dates)}" if dates else ""
    route = first["route"] or first["destination"] or first["origin"]
    case_count = len(rows)
    case_summary = (
        f"{first['maneuver_label']} | {first['vessel_type']} | {route or 'sem registo'} | "
        f"{first['loa_band']} | {case_count} caso(s) | rebocadores mais comuns {dominant_tug} | "
        f"dura\u00e7\u00e3o mediana {label_hours(duration_median)}"
    )
    plan_observations = " | ".join(comments[:3])
    vessel_loa = median([row["loa"] for row in rows])
    vessel_beam = median([row["beam"] for row in rows])
    vessel_gt = median([row["gt"] for row in rows])
    vessel_draft = median([row["draft"] for row in rows])
    source_rows = [str(row["no"]) for row in rows[:20]]
    feature = {
        "maneuver_type": first["maneuver_type"],
        "origin": first["origin"],
        "destination": first["destination"],
        "origin_key": key_text(first["origin"]),
        "destination_key": key_text(first["destination"]),
        "origin_is_anchorage": is_anchorage_berth(first["origin"]),
        "destination_is_anchorage": is_anchorage_berth(first["destination"]),
        "origin_is_known_berth": is_known_berth_label(first["origin"]),
        "destination_is_known_berth": is_known_berth_label(first["destination"]),
        "vessel_type": first["vessel_type"],
        "vessel_type_key": key_text(first["vessel_type"]),
        "vessel_loa_m": vessel_loa,
        "vessel_beam_m": vessel_beam,
        "vessel_gt_t": vessel_gt,
        "planned_draft_m": vessel_draft,
        "reported_draft_m": vessel_draft,
        "bow_thruster": "unknown",
        "stern_thruster": "unknown",
        "tug_count": dominant_tug,
        "constraints": [],
        "wave_sensitive": is_wave_sensitive(first["vessel_type"]),
    }
    return {
        "id": record_id,
        "maneuver_id": record_id,
        "port_call_id": "",
        "reference_code": f"EXP-{record_id[-6:].upper()}",
        "source_type": "practice_import",
        "source_label": "Experi\u00eancia pr\u00e1tica importada",
        "source_filename": "practice_maneuver_experience.json",
        "source_rows": source_rows,
        "vessel_name": f"Padr\u00e3o {first['vessel_type']} \u00b7 {route or 'sem registo'}",
        "maneuver_type": first["maneuver_type"],
        "maneuver_type_label": first["maneuver_label"],
        "current_state": "completed",
        "current_state_label": "Realizada",
        "origin_label": first["origin"],
        "destination_label": first["destination"],
        "planned_at": None,
        "decided_at": None,
        "completed_at": None,
        "reported_at": None,
        "latest_event_at": generated_at,
        "case_summary": case_summary,
        "practice_summary": case_summary,
        "practice_metrics": {
            "case_count": case_count,
            "date_range": date_range,
            "duration_median_h": duration_median,
            "duration_mean_h": duration_mean,
            "duration_median_label": label_hours(duration_median),
            "duration_mean_label": label_hours(duration_mean),
            "dominant_tug_count": dominant_tug,
            "tug_distribution_label": distribution_label(tug_values),
            "vessel_examples": examples,
            "comments_count": len(comments),
            "comments": comments[:5],
            "loa_band": first["loa_band"],
            "beam_band": first["beam_band"],
            "draft_band": first["draft_band"],
        },
        "vessel_snapshot": {
            "type": first["vessel_type"],
            "loa_m": format_number(vessel_loa),
            "beam_m": format_number(vessel_beam),
            "gt_t": format_number(vessel_gt),
            "max_draft_m": format_number(vessel_draft),
            "examples": examples,
        },
        "scale_snapshot": {
            "source": "practice_maneuver_experience.json",
            "date_range": date_range,
            "notes": "Dados agregados de pr\u00e1tica importada; a folha n\u00e3o inclui meteorologia nem IMO.",
        },
        "planning_snapshot": {
            "origin": first["origin"],
            "destination": first["destination"],
            "planned_draft_m": format_number(vessel_draft),
            "tug_count": dominant_tug,
            "plan_note": case_summary,
            "plan_observations": plan_observations,
            "created_by": "admin",
            "created_at": generated_at,
        },
        "decision_snapshot": {
            "decision": "approved",
            "state": "completed",
            "approval_note": "Padr\u00e3o agregado de experi\u00eancia pr\u00e1tica importada pelo admin.",
            "decided_by": "admin",
            "decided_at": generated_at,
        },
        "execution_snapshot": {
            "report_note": plan_observations,
            "reported_by": "experi\u00eancia importada",
            "reported_at": generated_at,
        },
        "outcome_snapshot": {
            "state": "completed",
            "state_label": "Realizada",
            "report_completed": True,
            "decision_flags": [],
        },
        "environment_snapshot": {
            "latest": {
                "status": "not_captured",
                "source": "practice_import",
                "reason": "A folha de pr\u00e1tica n\u00e3o inclui meteorologia nem leitura costeira.",
            }
        },
        "feature_snapshot": feature,
        "change_log": [],
        "feedback_status": "review",
        "feedback_note": "",
        "feedback_updated_by": "admin",
        "feedback_updated_at": generated_at,
        "created_at": generated_at,
        "updated_at": generated_at,
    }


def build_payload(rows: list[dict[str, Any]], skipped_rows: int, generated_at: str) -> dict[str, Any]:
    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[group_key(row)].append(row)
    records = [build_record(group, items, generated_at) for group, items in grouped.items()]
    records.sort(
        key=lambda record: (
            TYPE_ORDER.get(record["maneuver_type_label"], 99),
            -int(record["practice_metrics"]["case_count"]),
            record["case_summary"],
        )
    )
    type_counts = Counter(row["maneuver_label"] for row in rows)
    maneuver_types_label = ", ".join(
        f"{label} ({type_counts[label]})"
        for label in (LABEL_ENTRY, LABEL_DEPARTURE, LABEL_SHIFT, LABEL_ANCHOR)
        if type_counts[label]
    )
    return {
        "kind": KIND,
        "version": VERSION,
        "source": {
            "filename": "Manobras Pratica.xlsx",
            "worksheet": "Dados",
            "note": "Generated locally from the practice spreadsheet; pilot names are not stored.",
        },
        "generated_at": generated_at,
        "stats": {
            "source_filename": "practice_maneuver_experience.json",
            "raw_rows": len(rows),
            "skipped_rows": skipped_rows,
            "pattern_count": len(records),
            "comments_count": sum(record["practice_metrics"]["comments_count"] for record in records),
            "maneuver_types_label": maneuver_types_label,
        },
        "normalization": {
            "lisnave_aliases": "C3A/C3B map to Lisnave repair quays; D31/D32/D33 map to Lisnave dry docks.",
            "shift_route_rule": "For Mudanca records, Cais values with '/' are interpreted as origin/destination except known slash aliases such as Secil/Outao W.",
            "blank_tug_rule": "Rows with blank Rebocadores are treated as canceled/aborted and excluded; numeric 0 means zero tugs and is included.",
            "decimal_tug_rule": "Non-integer Rebocadores values are rounded to the nearest whole tug count.",
            "barge_dimension_rule": "Patterns like '3 x 29/18' are treated as three units, total LOA 87 m and beam 18 m.",
        },
        "records": records,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--sheet", default="Dados")
    parser.add_argument("--check", action="store_true", help="Generate and validate without writing.")
    args = parser.parse_args()

    rows, skipped_rows = parse_rows(args.input, args.sheet)
    generated_at = datetime.now(timezone.utc).isoformat()
    payload = build_payload(rows, skipped_rows, generated_at)
    if args.check:
        print(
            json.dumps(
                {
                    "records": len(payload["records"]),
                    "raw_rows": payload["stats"]["raw_rows"],
                    "skipped_rows": payload["stats"]["skipped_rows"],
                    "comments_count": payload["stats"]["comments_count"],
                    "maneuver_types_label": payload["stats"]["maneuver_types_label"],
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 0
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {args.output} with {len(payload['records'])} patterns from {len(rows)} rows.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
